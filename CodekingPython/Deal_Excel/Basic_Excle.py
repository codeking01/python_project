import openpyxl
from openpyxl import Workbook, load_workbook
from tqdm import tqdm


class ExcelData:
    def __init__(self, cas_list=[], deal_cas_list=[], rows=0):
        '''
        :param cas_list: 22_101_1的cas
        :param deal_cas_list: 22-101-1的cas
        :param rows:  excel的最大的列
        '''
        self.cas_list = cas_list
        self.deal_cas_list = deal_cas_list
        self.rows = rows


class DoExcel:
    def __init__(self, cas_list=[], deal_cas_list=[], read_path='', sheet_name='cas'):
        self.cas_list = cas_list
        self.deal_cas_list = deal_cas_list
        self.read_path = r'{read_path}'.format(read_path=read_path)
        self.sheet_name = sheet_name

    def Read_Data(self):
        '''
        这个方法是读取单列的CAS号的方法 后期需要多列再优化。
        :param read_path:  传入读取的excel的地址 精确到 xx.xlsx
        :param sheet_name:  传入读取的excel的sheet_name
        :return:  返回MyExcelData对象
        rows: 最大列，
        cas_list: 原始excel里面value，
        deal_cas_list: 返回20-101-1这种形式的cas
        '''
        if ('\\' in self.read_path):
            read_path = str(self.read_path).replace('\\', '/')
        else:
            read_path = str(self.read_path)
        sheet_name = self.sheet_name
        wb = openpyxl.load_workbook(read_path)
        sheet = wb[sheet_name]
        # 读取excel中的最大行数
        rows = sheet.max_row
        # 处理所有数据
        for sheet_row in tqdm(range(rows + 1)):
            try:
                # 读取cas号，从第二行开始
                cas = sheet.cell(row=sheet_row + 2, column=1).value
                deal_cas = cas.replace('_', '-')
                self.cas_list.append(str(cas))
                self.deal_cas_list.append(str(deal_cas))
            except:
                self.cas_list.append(str(cas))
                # 把原始的cas传进去
                self.deal_cas_list.append(str(cas))
                print('处理失败的{cas}'.format(cas=cas))
        # 把数据存储给对象
        MyExcelData = ExcelData(rows=rows, cas_list=self.cas_list, deal_cas_list=self.deal_cas_list)
        print("数据全部加载成功...")
        return MyExcelData


# 把xpath爬取的内容存储到excel里面
def WriteToExcle(path='', ExcelContent={}, CAS=None, current_index=None):
    path=r'{path}'.format(path=path)
    if ('\\' in path):
        save_path = str(path).replace('\\', '/')
    else:
        save_path = str(path)
    # 获取当前索引
    current_index = current_index + 1
    flag = 1
    try:
        # 写入excel中
        data = load_workbook(save_path)
    except:
        flag = 0
        print('excle没有新建哦，下次记得新建哈，这次我帮你了')
        wb = Workbook()
        data = wb.active
        data.title = 'cas'
        wb.save(save_path)
    finally:
        if (flag == 0):
            data = load_workbook(save_path)
        # print(data.defined_names.definedName) # 输出工作页索引范围
        # print(data.sheetnames) # 输出所有工作页的名称
        # 取第一张表
        sheetnames = data.sheetnames
        # 获取第一张表
        table = data[sheetnames[0]]
        # 这个是获取最后停留的excel的表
        # table = data.active
        # print(table.title) # 输出表名
        # nrows = table.max_row # 获得行数
        # ncolumns = table.max_column # 获得列数
        # 获取字典的所有键名
        xpath_key_list = []
        for j in ExcelContent:
            xpath_key_list.append(j)
        # 如果第一次存储数据，先写表头
        if (current_index == 1):
            # 在第一行写入对应的字段的名字，第一列默认写死，写成cas号
            table.cell(1, 1, 'CAS号')
            for i in range(len(xpath_key_list)):
                table.cell(1, i + 2, xpath_key_list[i])

        # 根据ExcleContent字典的长度，确定写入的列数
        for k in range(len(xpath_key_list)):
            if (k == 0):
                table.cell(current_index + 1, 1, CAS)
            # 从第二列开始写取出来的数据
            table.cell(current_index + 1, k + 2, ExcelContent[xpath_key_list[k]])
        # 存储成功
        data.save(save_path)

# 往已有的excle里面追加新的内容
# import openpyxl
# data = openpyxl.load_workbook(r'C:/Users/king/Desktop/test.xlsx')
# print(data.defined_names.definedName) # 输出工作页索引范围
# print(data.sheetnames) # 输出所有工作页的名称
# # 取第一张表
# sheetnames = data.sheetnames
# table=data[sheetnames[1]]
# # 过时了 table = data.get_sheet_by_name(sheetnames[0])
# table = data.active
# print(table.title) # 输出表名
# nrows = table.max_row # 获得行数
# ncolumns = table.max_column # 获得行数
# values = ['E','X','C','E','L']
# for value in values:
#     table.cell(nrows+1,1).value = value
#     nrows = nrows + 1
# data.save(r'C:/Users/king/Desktop/test.xlsx')
