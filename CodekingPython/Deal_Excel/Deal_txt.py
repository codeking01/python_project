import re

import requests
from openpyxl import Workbook
from tqdm import tqdm


def ReadTxt(path):
    a = ''
    if ('\\' in path):
        path = str(path).replace('\\', '/')
    else:
        path = str(path)
    with open(path, encoding='utf-8', mode='r') as f:
        lines = f.readlines()
    # 读取文件的内容并且存成一行字符串

    for line in lines:
        a += line.strip()
    # 设置一个列表去存储所有的数据
    b_list = a.split('第')
    b_list.remove(b_list[0])
    # 完整的列表数据
    cas_list = []
    for i in tqdm(range(len(b_list))):
        cas = b_list[i].split(':')[1]
        cas_list.append(cas)
        # print(cas)

    # 把失败的cas写入excle里面
    # 创建Excel文件
    wbk = Workbook()
    ws = wbk.active
    ws.title = 'cas'
    ws['A1'] = 'fail_cas'
    for i in tqdm(range(len(cas_list))):
        ws['A' + str(i + 2)].value = cas_list[i]
    wbk.save('mol失败的cas汇总.xlsx')
