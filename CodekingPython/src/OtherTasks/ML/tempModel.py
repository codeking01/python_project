"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/6/23 20:48
    @File : tempModel.py
"""
#%%
import pandas as pd
from numpy import unique

from src.OtherTasks.ML.PreDeal_Tools import Del_deletion_data, Record_usable_cols, Deal_sorted_Ydata, convert_to_num, \
    get_final_useablecols, all_ydata, delAndGetCols, get_merge_tabledata, deal_verify_ydata, get_former_Ydata
import numpy as np
# 这个pandas处理数据效果不太好 建议用numpy
thedata = pd.read_excel(io='Data613.xlsx', sheet_name='数据标准化-建模用')
Preddata = pd.read_excel(io='Data613.xlsx', sheet_name='数据标准化-预测用')
# 用来查看数据结果的excel对象
from openpyxl import load_workbook
filename='Data613.xlsx'
data = load_workbook(filename)
sheetnames = data.sheetnames
table = data[sheetnames[2]]

# 获取数据表对象
thedata = np.array(thedata)
Preddata = np.array(Preddata)
# 将表的内容 7个数据表 挨个处理
#%%
# 获取源建模的X数据 传入的对象是获取的excel对象
FormerTwo_data, FormerThree_data, FormerFour_data, FormerFive_data, FormerSix_data, FormerSeven_data = get_merge_tabledata(thedata)
# 获取预测的X数据
Verify_TABLE_TWO, Verify_TABLE_THREE, Verify_TABLE_FOUR, Verify_TABLE_FIVE, Verify_TABLE_SIX, Verify_TABLE_SEVEN = get_merge_tabledata(Preddata)

# 删除缺失值过多的列，并保存del_cols
# tempFormerTwo_data = delAndGetCols(FormerTwo_data)
base_Xdata,del_cols = delAndGetCols(FormerSeven_data)
# 用del_cols删除验证集X不需要的列
Verify_Xdata = np.delete(Verify_TABLE_SEVEN, del_cols, axis=1)

# base_Xdata,del_cols
#%%
# 获取源建模的数据Y_data
Original_TableTwoYdata, Original_TableThreeYdata, Original_TableFiveYdata, Original_TableSixYdata, Original_TableSevenYdata=get_former_Ydata(thedata)
# 获取预测数据的Y_data
Predict_TableTwoYdata, Predict_TableThreeYdata, Predict_TableFiveYdata, Predict_TableSixYdata, Predict_TableSevenYdata=get_former_Ydata(Preddata)
#%%
# 将Y取出来，然后取出缺失值过多的列
# 获取可用的列
final_cols = get_final_useablecols(Original_TableSevenYdata, Predict_TableSevenYdata)
# final_cols
#%%
# 获取可以用的Y_data
Original_YdataList, Pred_YdataList = all_ydata(final_cols, Original_TableSevenYdata, Predict_TableSevenYdata)
#%%
# 挨个遍历，取出有用的Y_data
# 传入起始的坐标 24,67,108,145,179
the_column=179
start=0
for index in final_cols:
    # 建模的Y_data
    Modeling_Y_data = Original_YdataList[start]
    # 预测的Y_data
    Pred_Y_data = Pred_YdataList[start]
    start += 1
    # 处理建模的X_data,Y_data
    temp_OriginalXdata = np.column_stack((base_Xdata, Modeling_Y_data))
    # 删除缺失的行
    temp_OriginalXdata = Del_deletion_data(temp_OriginalXdata, 0)
    X_data = temp_OriginalXdata[:, :-1]
    Y_data = temp_OriginalXdata[:, -1]
    # 处理验证的X_data,Y_data
    # 合并数据
    temp_verifydata = np.column_stack((Verify_Xdata, Pred_Y_data))
    # 删除缺失的行
    temp_verifydata = Del_deletion_data(temp_verifydata, 0)
    # 获取预测数据的X_data，Y_data
    X_verify_data = temp_verifydata[:, :-1]
    Y_verify_data = temp_verifydata[:, -1]
    print(list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist())), list(set(np.where(np.isnan(temp_verifydata.astype(float)) == True)[0].tolist())))
    # 将建模Y_data分类(-1,0,1)，并且取出边界值
    Y_data,Y_data_boundsMin,Y_data_boundsMax=Deal_sorted_Ydata(Y_data)
    # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
    Verify_Ydata = deal_verify_ydata(Y_verify_data, Y_data_boundsMin, Y_data_boundsMax)
    # 查看是否替换成功
    print([unique(Y_data),Y_data_boundsMin,Y_data_boundsMax],np.unique(Verify_Ydata))
    # 切分训练数据和测试数据
    from sklearn.model_selection import train_test_split
    ## 30%测试数据，70%训练数据，stratify=y表示训练数据和测试数据具有相同的类别比例
    X_train, X_test, y_train, y_test = train_test_split(X_data, Y_data, test_size=0.3, random_state=2, stratify=Y_data)
    from sklearn.metrics import accuracy_score
    # 开始训练模型
    from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
    # 获取模型  随机森林
    Rfc = RandomForestClassifier()
    Rfc.fit(X_train, y_train.astype(float))
    # 训练集准确率
    Rfc_train_acc =accuracy_score(y_train.astype(float), Rfc.predict(X_train))
    print('Rfc_train_acc准确率：', Rfc_train_acc)
    Rfc_pred = Rfc.predict(X_test)
    Rfc_acc = accuracy_score(y_test.astype(float), Rfc_pred)
    # 测试集准确率
    print('Rfc_acc准确率:', Rfc_acc)

    # 梯度提升树
    Gbc = GradientBoostingClassifier()
    Gbc.fit(X_train, y_train.astype(float))
    # 训练集准确率
    Gbc_train_acc =accuracy_score(y_train.astype(float), Gbc.predict(X_train))
    print('Gbc_train_acc准确率：', Gbc_train_acc)
    # 测试集准确率
    Gbc_pred = Gbc.predict(X_test)
    Gbc_acc = accuracy_score(y_test.astype(float), Gbc_pred)
    print('Gbc_acc准确率:', Gbc_acc)

    # 使用SVM模型
    from sklearn.svm import SVC
    Svc = SVC()
    Svc.fit(X_train, y_train.astype(float))
    # 训练集准确率
    Svc_train_acc =accuracy_score(y_train.astype(float), Svc.predict(X_train))
    print('Svc_train_acc准确率：', Svc_train_acc)
    # 测试集准确率
    Svc_pred = Svc.predict(X_test)
    Svc_acc = accuracy_score(y_test.astype(float), Svc_pred)
    print('SVC准确率：', Svc_acc)

    # 使用DNN模型
    from sklearn.neural_network import MLPClassifier
    Dnn = MLPClassifier()
    Dnn.fit(X_train, y_train.astype(float))
    # 训练集准确率
    Dnn_train_acc =accuracy_score(y_train.astype(float), Dnn.predict(X_train))
    print('Dnn_train_acc准确率：', Dnn_train_acc)
    # 测试集准确率
    Dnn_pred = Dnn.predict(X_test)
    Dnn_acc = accuracy_score(y_test.astype(float), Dnn_pred)
    print('DNN准确率：', Dnn_acc)

    # 使用卷积神经网络模型
    from sklearn.neural_network import MLPClassifier
    CNN = MLPClassifier()
    CNN.fit(X_train, y_train.astype(float))
    # 训练集准确率
    CNN_train_acc =accuracy_score(y_train.astype(float), CNN.predict(X_train))
    print('CNN_train_acc准确率：', CNN_train_acc)
    # 测试集准确率
    CNN_pred = CNN.predict(X_test)
    CNN_acc = accuracy_score(y_test.astype(float), CNN_pred)
    print('CNN准确率：', CNN_acc)

    # 预测数据
    Rfc_pred = Rfc.predict(X_verify_data)
    Gbc_pred = Gbc.predict(X_verify_data)
    Svc_pred = Svc.predict(X_verify_data)
    Dnn_pred = Dnn.predict(X_verify_data)
    CNN_pred = CNN.predict(X_verify_data)
    # 计算准确率
    verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)
    verify_y2 = accuracy_score(Y_verify_data.astype(float), Gbc_pred)
    verify_y3 = accuracy_score(Y_verify_data.astype(float), Svc_pred)
    verify_y6 = accuracy_score(Y_verify_data.astype(float), Dnn_pred)
    verify_y7 = accuracy_score(Y_verify_data.astype(float), CNN_pred)

    # 计算准确率
    print('验证集Rfc_acc准确率:', verify_y1)
    print('验证集Gbc_acc准确率:', verify_y2)
    print('验证集SVC准确率：', verify_y3)
    print('验证集DNN准确率：', verify_y6)
    print('验证集CNN准确率：', verify_y7)

    # 将训练集写入到excel中
    table.cell(5, the_column+index).value = Rfc_train_acc
    table.cell(6, the_column+index).value = Gbc_train_acc
    table.cell(7, the_column+index).value = Svc_train_acc
    table.cell(8, the_column+index).value = Dnn_train_acc
    table.cell(9, the_column+index).value = CNN_train_acc

    # 将测试集写入到excel中
    table.cell(11, the_column+index).value = Rfc_acc
    table.cell(12, the_column+index).value = Gbc_acc
    table.cell(13, the_column+index).value = Svc_acc
    table.cell(14, the_column+index).value = Dnn_acc
    table.cell(15, the_column+index).value = CNN_acc

    # 将预测数据写入到excel中
    table.cell(17, the_column+index).value = verify_y1
    table.cell(18, the_column+index).value = verify_y2
    table.cell(19, the_column+index).value = verify_y3
    table.cell(20, the_column+index).value = verify_y6
    table.cell(21, the_column+index).value = verify_y7

    # 保存excel文件
    data.save(filename)
