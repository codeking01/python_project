# auth: code_king
# time: 2022/7/10 23:44
# file: properties.py
"""
    起始网址： https://polymer.nims.go.jp/PoLyInfo/cgi-bin/p-search.cgi
    判断标签samples是否小于等于30
"""
import re
import time

import requests
import winsound
from lxml import etree
from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

cookie = 'MOD_AUTH_CAS=ec9b1e01daff6314027b750cb3d41047; TS0164a68e=01c6fd78ba1e62b1f2472a31dab99abb6986430c50395cd733274e83896ad0b6e17e3378bc6d4379d6090e4481f9a6cfd31d7ff59d4a78ae9fe43278dc7ccb543c9090067c; _ga=GA1.3.543165510.1657631087; _gid=GA1.3.76842478.1657631087; TS01ec3135=01c6fd78ba095010781a7dfad6df403d5e51614526fda064a89f232d94949201dab32c7876749e92cd1ed121ea20ad744a114e345a; TS7d5e355c_76=0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f208846012bd09e80088dcd9464a427752dd3d42dcb52cf8b4b9f958b603a5128833c7336e1b98f2f16d0a8d8c357a95d41f9db7322e749bd91bbc0ff32671cdc3baf3e541c01727ef0f570ea6b3942e8f33df57e64d383e45c8897141d2b7e1c99701f7a0fecd0f02fe3decbdd08a1d2c5d784ab61df27dfd23f6e3111d56b76a5eab6d539e5d07762c474f821ae31c080abeed16b6e657f58311373c50dc4f3fd882c2cb137d4dc8525c619e64275bb19821ccf3f2c9d277decb4a9f600882a0801c4c1f5541254c85168cc2872d59ef01425e2f5f4f50c9c6c96834be744e0de5d915d773e536231352c47fc0ce5db8; TSPD_101=0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f2:0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f208846012bd063800a3f8a1c567cb2eb17dcf8599b25a11a02edac807bd6de97e02ad698ef45a600827c3e7573a7884ba4a4c08b9142bac4f5f94900f8370829f; _gat_gtag_UA_162281910_1=1; TSbe9bed7a_27=0821037767ab2000dcce8ab2b2bd6615b71d009d896c4ae8dbe4ef22241f5157d5bff32916b3571b08231829fa11200088a4b04cb9c5b5883baabeaa1a99531fd4439533a178bd1a3618e1da428f2b5e'
sec_cookies = 'MOD_AUTH_CAS=ec9b1e01daff6314027b750cb3d41047; TS0164a68e=01c6fd78ba1e62b1f2472a31dab99abb6986430c50395cd733274e83896ad0b6e17e3378bc6d4379d6090e4481f9a6cfd31d7ff59d4a78ae9fe43278dc7ccb543c9090067c; _ga=GA1.3.543165510.1657631087; _gid=GA1.3.76842478.1657631087; TS01ec3135=01c6fd78ba095010781a7dfad6df403d5e51614526fda064a89f232d94949201dab32c7876749e92cd1ed121ea20ad744a114e345a; TS7d5e355c_76=0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f208846012bd09e80088dcd9464a427752dd3d42dcb52cf8b4b9f958b603a5128833c7336e1b98f2f16d0a8d8c357a95d41f9db7322e749bd91bbc0ff32671cdc3baf3e541c01727ef0f570ea6b3942e8f33df57e64d383e45c8897141d2b7e1c99701f7a0fecd0f02fe3decbdd08a1d2c5d784ab61df27dfd23f6e3111d56b76a5eab6d539e5d07762c474f821ae31c080abeed16b6e657f58311373c50dc4f3fd882c2cb137d4dc8525c619e64275bb19821ccf3f2c9d277decb4a9f600882a0801c4c1f5541254c85168cc2872d59ef01425e2f5f4f50c9c6c96834be744e0de5d915d773e536231352c47fc0ce5db8; TSPD_101=0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f2:0821037767ab2800a4ceea251252caddbb2301199b41e882e5ef1228e01a620da122447ee45347c1bae57033beb771f208846012bd063800a3f8a1c567cb2eb17dcf8599b25a11a02edac807bd6de97e02ad698ef45a600827c3e7573a7884ba4a4c08b9142bac4f5f94900f8370829f; TSbe9bed7a_27=0821037767ab20007786fecc4f4cd70b843df8cf9775b1f2fb494c7999d23ec24a40622905a3505f08a1d0034d112000c6db4aa119e898476e2174346bd8d9e4fe6f1f3cd4f3641c97aa8c987cd1915b'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    'cookie': cookie,
}
second_headers={
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    'cookie': sec_cookies,
}
# "https://polymer.nims.go.jp/PoLyInfo/cgi-bin/ho-id-search.cgi?PID=Yd6l2119R16f9P19gshngrlq5ioiqlqiomo7P&SID=gsktommpmo5-fpjo3-hqkpgn5-epe2&layout=info"
#                                             "ho-id-search.cgi?PID=Yd6l2119R16f9P19gshngrlq5ioiqlqiomo7P&SID=gsktommpmo5-fpjo3-hqkpgn5-epe2&layout=info"

def save_urls(download_url):
    wb = load_workbook('download_urls.xlsx')
    sheetnames = wb.sheetnames
    # 获取第一张表
    sheet = wb[sheetnames[0]]
    # 计算最大行数
    max_row = sheet.max_row
    sheet.cell(max_row+1, 1).value = download_url
    # 每500个保存一次
    wb.save('download_urls.xlsx')


def get_store_url(PID, SID, c):
    original_url = 'https://polymer.nims.go.jp/PoLyInfo/cgi-bin/ho-id-search.cgi'
    data = {
        'layout': 'sample',
        'PID': PID,
        'SID': SID,
        'c': c
    }
    response = requests.post(original_url, data=data, headers=second_headers)
    tree = etree.HTML(response.text)
    # //table[2]//tr[n]/td[2]/a/@href
    for index in range(2, 31):
        url = tree.xpath(f'//table[2]//tr[{index}]/td[2]/a/@href')
        # 默认最大的样品数量为30，所以只需要挨个遍历就可以了
        if index == 2:
            if  len(url) == 0:
                print('更换cookie!')
                winsound.Beep(3000, 3000)
                time.sleep(10)
        if len(url) != 0:
            download_url = 'https://polymer.nims.go.jp/PoLyInfo/cgi-bin/' + url[0]
            # 下载写入到excel中
            save_urls(download_url)
        else:
            break


if __name__ == '__main__':
    base_url = 'https://polymer.nims.go.jp/PoLyInfo/cgi-bin/p-search.cgi'
    pagesize = 1800
    data = {
        'pttab': 'Homopolymer',
        'search': 'page',
        'PIDs': 12807,
        'rPIDs': '',
        'pagesize': pagesize,
    }
    response = requests.post(base_url, data=data, headers=headers)
    # requests中的session会帮我们得到cookies的值
    # session = requests.session()
    # response = session.post(base_url, data=data, headers=headers)  # response 响应对象
    # res = session.get(base_url, headers=headers).content.decode()

    if response.status_code == 200:
        res_content = response.text
        # 使用xpath解析
        tree = etree.HTML(res_content)
        for index in tqdm(range(1, pagesize)):
            try:
                # 获取的是一个列表，先获取第一个元素，并且判断这个xx sample前面的数字是否小于30
                input_value = tree.xpath(f"//form[{index}]//table[1]//@value")
                # 解析结束标志
                if len(input_value) == 0:
                    break
                # 使用正则获取里面的数字，并且判断数字是否小于等于30
                num_pattern = re.compile(r"\d+")
                num = int(num_pattern.findall(input_value[0])[0])
                # 如果小于等于30，则存储数据
                if num <= 30:
                    # 获取标签内容
                    PID, SID, c = input_value[2], input_value[3], input_value[4]
                    # 去获取sample的信息
                    get_store_url(PID, SID, c)
            except Exception as e:
                print(e)
                # 播放声音并且停顿10s
                winsound.Beep(3000, 3000)
                time.sleep(10)
                # 更换header(带cookie)
        print('感谢使用！')