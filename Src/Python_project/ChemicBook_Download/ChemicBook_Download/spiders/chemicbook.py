import openpyxl
import scrapy
from tqdm import tqdm

from ChemicBook_Download.items import ChemicbookDownloadItem

global cas, deal_cas, obj_cid, current_num, rows, headers,font_control
# 用来表示 cas和处理了-后的deal_cas : 分别格式为 100-2-201  100_-2_201
cas_list = []
deal_cas_list = []
# 用来存储下载链接
mol_url_list=[]
gif_url_list=[]

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
}

class FontColor:
    OKBULE = '\033[94m'
    OKGREEN = '\033[92m'
    WARN = '\033[93m'
    FAIL = '\033[91m'
font_control= FontColor()

# 处理最开的excel的数据
def Predeal_Data():
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['cas']
    # cas_list=[]
    # deal_cas_list=[]

    # 读取excel中的最大行数
    rows = sheet.max_row
    # 处理所有数据
    for sheet_row in tqdm(range(rows + 1)):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=sheet_row + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
        except:
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
            print(cas, deal_cas)
    print("数据全部加载成功...")
    return rows

class ChemicbookSpider(scrapy.Spider):
    name = 'chemicbook'
    allowed_domains = ['www.chemicalbook.com']
    # start_urls = ['http://www.chemicalbook.com/']
    # 重写start_requests
    def start_requests(self):
        url='http://www.chemicalbook.com/'
        print("重写start_requests")
        headers = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
        }
        self.headers=headers
        yield scrapy.Request(url=url,headers=self.headers, callback=self.parse)

    # mol下载地址 https://www.chemicalbook.com/CAS/MOL/100-02-7.mol
    # gif https://www.chemicalbook.com/CAS/GIF/100-02-7.gif
    def parse(self, response):
        # 获取总数据量
        rows = Predeal_Data()
        print("数据加载成功，开始爬~~~~~~~~")
        # 把cas号和下载链接一次性全部传到管道 然后再开始去判断
        for current_num in tqdm(range(0, rows+1)):
            deal_cas = str(deal_cas_list[current_num])
            # 后面下载失败自己校验
            mol_url=' https://www.chemicalbook.com/CAS/MOL/{mol_url}.mol'.format(mol_url=deal_cas)
            gif_url='https://www.chemicalbook.com/CAS/GIF/{png_url}.gif'.format(png_url=deal_cas)
            # 记录存储的下载链接
            mol_url_list.append(mol_url)
            gif_url_list.append(gif_url)
        print("下载链接全部都存储完毕！")
        All_Data=ChemicbookDownloadItem(mol_url_list=mol_url_list,gif_url_list=gif_url_list)
        yield All_Data
