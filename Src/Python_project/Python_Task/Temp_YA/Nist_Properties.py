"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/5/12 16:38
    @File : Nist_Properties.py
"""
import time

import requests
from lxml import etree
from openpyxl import load_workbook
from tqdm import tqdm


# 根据传来的字典秀娥如到excel里面
def save_to_excel(cas, current_index, table, content_dict):
    for key in content_dict:
        the_values = content_dict.get(key)
        # print(the_values)
        table.cell(current_index + 1, 1, cas)
        table.cell(current_index + 1, 2, str(key))
        table.cell(current_index + 1, 3, str(the_values))


def GetCasList(rows):
    casList = []
    for sheet_row in tqdm(range(rows + 1)):
        cas = sheet.cell(row=sheet_row + 2, column=1).value
        if (cas != None):
            casList.append(cas)
    return casList


if __name__ == '__main__':
    # 存储到失败的cas的号
    save_failpath = 'failcas.xlsx'
    cas_data = load_workbook(save_failpath)
    fail_sheetnames = cas_data.sheetnames
    cas_table = cas_data[fail_sheetnames[0]]
    # 失败的cas 的excel表
    fail_cas_index = 1

    # 存储的excel文件
    save_path = 'Nist_Properties.xlsx'
    data = load_workbook(save_path)
    sheetnames = data.sheetnames
    # 获取第一张表
    table = data[sheetnames[0]]
    #  读取excel的cas号
    read_path = 'test.xlsx'
    wb = load_workbook(read_path)
    sheet = wb['Sheet1']
    # 读取excel中的最大行数
    rows = sheet.max_row
    allcas_list = GetCasList(rows)
    current_index = 0
    # 挨个 遍历
    for index in tqdm(range(rows)):
        cas = allcas_list[index]
        # 取访问的时候的cas号
        search_cas=cas.replace('_','-')
        url = 'https://webbook.nist.gov/cgi/cbook.cgi?ID={cas}&Units=SI'.format(cas=search_cas)
        # 计算是否断网
        try:
            response = requests.get(url)
            res_content = response.text
            # next_url = tree.xpath(r'''//a[contains(text(),'NIST / TRC Web Thermo Tables, professional edition')]/@href''')
            # 看看是否有链接存在
            content_flag = 'Data at NIST subscription sites'

            # 判断有链接才接着往下走
            if content_flag in res_content:
                try:
                    tree = etree.HTML(res_content)
                    next_url = tree.xpath(r'//*[@id="main"]/ul[2]/li/a/@href')
                    next_url = next_url[0]
                    # 处理数据
                    next_content = requests.get(next_url).text
                    next_tree = etree.HTML(next_content)
                    i = 1
                    # 获取所有的性质
                    while (True):
                        # 用来存储性质内容的字典
                        content_dict = {}
                        eachproperty = next_tree.xpath(f'//div/ul/li[{i}]/text()')
                        if (len(eachproperty) == 0):
                            # 出现了空列表就结束循环
                            break
                        # 先判断是不是多标题的
                        eachproperties = next_tree.xpath(f'//div/ul/li[{i}]//li//text()')
                        if (len(eachproperties) == 0):
                            # 处理单标题
                            signal_propety = next_tree.xpath(f'//div/ul/li[{i}]//text()')
                            content_dict.update({str(signal_propety[0].strip()): str(signal_propety[1].strip())})
                        else:
                            # 存储多标题的
                            propetyname = next_tree.xpath(f'//div/ul/li[{i}]//text()')[0].strip()
                            eachproperties = [str(temp).strip() for temp in eachproperties]
                            propetyvalue = eachproperties
                            content_dict.update({str(propetyname): propetyvalue})
                        # 存储到excel中
                        # cas = cas.replace('-', '_')
                        cas = str(cas)
                        save_to_excel(cas, current_index, table, content_dict)
                        if(index%500==0):
                            data.save(save_path)
                        current_index += 1
                        i += 1
                        # print(f'第{index}个数据处理结束！cas号为{cas}')

                except Exception as e:
                    print(f'未知错误: {e}')
                    print(f'原始url: {url},最终url:{next_url}')
                    current_index += 1
                    data.save(save_path)
                    # 失败的存储进去
                    cas_table.cell(fail_cas_index + 1, 1, str(cas))
                    fail_cas_index += 1
                    cas_data.save(save_failpath)

            # 如果没有链接，则写入提示数据到excel里面
            else:
                cas = str(cas)
                # 只写cas号即可
                table.cell(current_index + 1, 1, cas)
                table.cell(current_index + 1, 2, '没有数据！')
                print(f'数据不存在，当前cas为{cas},链接为: {url}')
                data.save(save_path)
                current_index += 1
        except Exception as e:
            print('应该断网了-----------------------------------------------------')
            print(f'未知错误: {e}')
            print(f'cas: {cas}')
            current_index += 1
            data.save(save_path)
            # 失败的存储进去
            cas_table.cell(fail_cas_index + 1, 1, str(cas))
            fail_cas_index += 1
            cas_data.save(save_failpath)
            break
    print('整理完毕-----------')
    # 确保保存完毕
    data.save(save_path)
    cas_data.save(save_failpath)
