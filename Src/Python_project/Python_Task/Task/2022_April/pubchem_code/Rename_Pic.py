"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/4/17 11:13
    @File : Rename_Pic.py
"""
import os
import re
import shutil
import time
import openpyxl
from tqdm import tqdm


def get_cid_dic(read_path):
    try:
        # 读取excel的某列，全部读取出来，然后根据正则去取出cas号
        # read_path = '1_1w.xlsx'
        wb = openpyxl.load_workbook(read_path)
    except Exception as e:
        print("路径可能有问题，请检查,错误原因：", e)
    else:
        # 取第一张表名
        sheetnames = wb.sheetnames
        # 获取第一张表
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        max_column = ws.max_column
        cid_dic = {}
        for index in range(2, rows + 1):
            key = ws.cell(index, 2).value
            value = ws.cell(index, max_column).value
            cid_dic.update({str(key): str(value)})
        return cid_dic

if __name__ == '__main__':
    # 用来判断程序消耗时间
    t1 = time.time()
    # 读取excel的地址，会根据里面的cid去命名图片
    read_path = '1_1w.xlsx'
    # 读取excel
    cid_dic = get_cid_dic(read_path)
    # 图片所在文件夹，注意格式的书写
    filePath = 'D:\\test1'
    name_list = os.listdir(filePath)
    cas_pattern = re.compile(r"\d{2,7}_\d{2}_\d{1}")
    for i in tqdm(range(0, len(name_list))):
        # 原名
        src = filePath + '\\' + name_list[i]
        # todo 有可能cid对应的cas是空的
        try:
            try:
                # 找到cid
                cid_key = name_list[i].split('_')[1].split('.')[0]
                # 当cid不存在时,跳出循环
                if(cid_key not in cid_dic):
                    casname=''
                    print('cid_key有问题,excel中不存在这个数，原因应该是图片存在cid,cid为:{cid}'.format(cid=cid_key))
                    # todo 删除图片
                    os.remove(src)
                    continue
                else:
                    casname = cid_dic[cid_key]
            except Exception as e:
                print('未知错误,失败原因:{e}'.format(e=e))
                casname=''

            #用正则取出cas号存储到列表里
            casname_list = cas_pattern.findall(casname)
            # casname_list为空，则没有cas号
            if (len(casname_list) == 0):
                # cas不存在的话，直接删除这个cid对应的图片
                os.remove(src)
                print('cas不存在，已经删除cid的原图了,cid为{cid}'.format(cid=name_list[i]))
                print('第{i}条数据删除成功。'.format(i=i))
                continue
            #  cid对应多个cas则多存储几次
            elif (len(casname_list) > 1):
                try:
                    # 依次取出cas号，并且做重命名
                    for index in range(len(casname_list)):
                        new_name = casname_list[index]
                        new_file_name = filePath + '\\' + str(new_name)+'.png'
                        shutil.copyfile('{old_file}'.format(old_file=src), '{new_file}'.format(new_file=new_file_name))
                    # 删除源文件
                    os.remove(src)
                except Exception as e:
                    print('文件名可能有误，失败原因：{e}'.format(e=e))
            # 只有一个的情况的下，只需要重名即可
            elif (len(casname_list) ==1):
                dst=filePath + '\\'+casname_list[0]+'.png'
                os.rename(src, dst)
            print('第{i}条数据转化成功。'.format(i=i))
        except Exception as e:
            print('未知错误，reasons: {e}'.format(e=e))
    t2 = time.time()
    t=round(t2 - t1, 2)
    print('程序运行结束，感谢您的使用，本次消耗时间为：{t}'.format(t=t))