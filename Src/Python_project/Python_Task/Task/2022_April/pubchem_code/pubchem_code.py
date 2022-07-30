"""
    -- coding: utf-8 --
    @Author: codeking
    @Data : 2022/4/15 14:42
    @File : pubchem_code.py
"""
import time
import re
from time import sleep
import openpyxl
from pandas import read_csv
import os


# 自定义路径的文件创建
from tqdm import tqdm


def creat_save_path(filename, load_path, default_flag):
    try:
        if (default_flag == True):
            dir_path = r'{load_path}'.format(load_path=load_path)
            os.mkdir(dir_path + './' + str(filename))
        else:
            # 自定义保存的位置
            dir_path = r'{load_path}'.format(load_path=load_path)
            os.mkdir(dir_path + '\\' + str(filename))
            dir_path = dir_path.replace('\\', '/')
            return dir_path
    except Exception as e:
        print('取路径可能有误，或者文件已经存在读(程序没停止就没关系，忽略即可),', e)


# 存储cas号到excel中
def get_cas_list(excel_path, turn_flag):
    try:
        # 读取excel的某列，全部读取出来，然后根据正则去取出cas号
        read_path = excel_path
        wb = openpyxl.load_workbook(read_path)
    except Exception as e:
        print('路径可能有问题，请检查,错误原因：', e)
    else:
        # 取第一张表名
        sheetnames = wb.sheetnames
        # 获取第一张表
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        max_column = ws.max_column
        # 获取所有cas的数据 用cas_list存储（列表）
        cas_list = []
        # 用来存储cid数据
        cid_list = []

        # 定义正则规则
        cas_pattern = re.compile(r"\|(\d{2,7}-\d{2}-\d{1})\||\|CAS-(\d{2,7}-\d{2}-\d{1})\|")
        # 存储cas到excel里面 cas数据在第4列
        for index in range(2, rows + 1):
            pre_cas = ws.cell(row=index, column=4).value
            # 将cid号全部取出来 存在一个列表里 cid数据在第2列
            cid_list.append(ws.cell(row=index, column=2).value)

            # 部分的数据在excel里面没有
            if (pre_cas == None):
                cas_list.append('')
                ws.cell(index, max_column + 1, '')
            else:
                cas = cas_pattern.findall(pre_cas)
                # 考虑没有cas号的情况 存一个 '' 占位，确保顺序一致性，后面加判断
                if (len(cas) == 0):
                    cas_list.append('')
                    ws.cell(index, max_column + 1, '')
                else:
                    # 处理匹配到的数据  这个是这个类型  cas=[('71-55-6', ''), ('', '71-55-6')]
                    cas_temp_list = []
                    for cas_value_index in range(len(cas)):
                        cas_temp = cas[cas_value_index]
                        for c_value in cas_temp:
                            if (c_value != ''):
                                cas_temp_list.append(c_value)
                    # 删除重复的cas数据
                    cas = list(set(cas_temp_list))
                    # 在有cas的情况下，分别考虑多个cas和单个cas的情况
                    # 将'-'转换为'_'
                    cas = [i.replace('-', '_') for i in cas]
                    # 如果取到多个cas号 则以列表的形式存储  把列表的所有数据存储进去
                    if (len(cas) > 1):
                        temp_list = []
                        for storage_cas in cas:
                            # # 判断0开头的cas号，不是0才存进去
                            # if (storage_cas[0] != '0'):
                            temp_list.append(storage_cas)
                        cas_list.append(temp_list)
                        ws.cell(index, max_column + 1, '{data}'.format(data=temp_list))
                    else:
                        # cas列表可能为空
                        if (len(cas) == 0):
                            cas_list.append('')
                            ws.cell(index, max_column + 1, '')
                        else:
                            cas_list.append(cas[0])
                            ws.cell(index, max_column + 1, '{data}'.format(data=cas[0]))
        if (turn_flag == True):
            wb.save(excel_path)
        print("excel的cas全部存储完毕！")
        return cas_list, cid_list


def Read_SDF(read_path):
    try:
        # 读取存储的sdf文件
        with open('{read_path}'.format(read_path=read_path), 'r', encoding='utf-8') as sdf_file:
            sdf_file = sdf_file.readlines()
            print('sdf文件读取成功')
            return sdf_file
    except Exception as e:
        print('sdf文件有问题，请检查一下，失败原因：', e)


# 把数据写入为mol，并且以cas号命名
def save_mol(savefilepath, data, temp_list):
    try:
        # todo 判断一下存储的cas是否原来也存储过
        # 处理cas是列表的数据
        if (type(data) == list):
            for L_index in data:
                # 判断一下存储的cas是否原来也存储过
                if(L_index in saved_caslist):
                    same_index=0
                    # 判断出现过几次了
                    for L in  saved_caslist:
                        if(L_index == L):
                            same_index+=1
                    # 修改原来的名字为 100_10_1-1
                    cas_name=L_index+'-'+str(same_index)
                    file = open('{savefile}/{cas}.mol'.format(savefile=savefilepath, cas=cas_name), 'w', encoding='utf-8')
                    for i in temp_list:
                        file.write(i)
                    file.close()
                    saved_caslist.append(data)
                else:
                    file = open('{savefile}/{cas}.mol'.format(savefile=savefilepath, cas=L_index), 'w', encoding='utf-8')
                    for i in temp_list:
                        file.write(i)
                    file.close()
                    saved_caslist.append(L_index)
        # 单个数据
        else:
            # 判断一下存储的cas是否原来也存储过
            if(data in saved_caslist):
                same_index=0
                # 判断出现过几次了
                for L in  saved_caslist:
                    if(data == L):
                        same_index+=1
                # 修改原来的名字为 100_10_1-1
                cas_name=data+'-'+str(same_index)
                file = open('{savefile}/{cas}.mol'.format(savefile=savefilepath, cas=cas_name), 'w', encoding='utf-8')
                for i in temp_list:
                    file.write(i)
                file.close()
                saved_caslist.append(data)
            else:
                file = open('{savefile}/{cas}.mol'.format(savefile=savefilepath, cas=data), 'w', encoding='utf-8')
                for i in temp_list:
                    file.write(i)
                file.close()
                saved_caslist.append(data)
    except Exception as e:
        # 考虑将失败的数据存入日志文件中
        with open('./log/{cas}.log'.format(cas='fail_sdf'), encoding='utf-8', mode='a') as file:
            # 将转化失败的存入日志中
            file.write(str(cas_list[current_index]))
            file.write('\n**失败原因：' + str(e) + '\n')
            file.close()
        print('\n****cas名称不合法，或者保存路径有误，请检查，失败原因{e}\n'.format(e=e))

# 判断cas是否为空
def IsNotEmpty(func):
    def wrapper(*args):
        if (str(sdf_cid) in cid_dic):
            # cas不为空的时候进行存储
            if (cid_dic[str(sdf_cid)] != ''):
                func(*args)
                print('第{num}数据转化成功'.format(num=current_index))
            # 处理cas为空的情况
            else:
                print('cas号不存在!')
        else:
            with open('./log/{cas}.log'.format(cas='fail_sdf'), encoding='utf-8', mode='a') as file:
                # 将转化失败的存入日志中
                file.write(str(sdf_cid))
                file.write('\n**失败原因: cid不在excel表里面！' + '\n')
                file.close()
            print('cid号不在当前字典中!')
    return wrapper


# 将sdf内容转化为mol存储，需要判断cas是否存在，如不存在则跳过
@IsNotEmpty
def convert_mol(savepath, temp_list, current_index, sdf_cid):
    data = cid_dic[str(sdf_cid)]
    save_mol(savefilepath=savepath, data=data, temp_list=temp_list)


#  由于csv文件操作存在一些问题，这里我转化为xlsx进行操作
def ConvertToExcel(path_name):
    csv_path = '{name}'.format(name=str(path_name))
    f = open(csv_path, encoding='utf-8')
    data = read_csv(f)
    path_name = path_name.split('.')[0]
    excel_path = '{path_name}.xlsx'.format(path_name=str(path_name))
    data.to_excel(excel_path)


# 获取cid对应的cas，将其存储为字典
def get_cid_dic(cid_list, cas_list):
    # 将cid数据 和 cas号对应起来
    cid_dic = {}
    if (len(cid_list) == len(cas_list)):
        for index in range(0, len(cas_list)):
            key = cid_list[index]
            value = cas_list[index]
            cid_dic.update({str(key): value})
        return cid_dic
    else:
        print('数据获取有误')


if __name__ == '__main__':
    # todo 将需要修改的东西放到配置文件
    t1 = time.time()
    # 输入csv路径
    path_name = '1_1w.csv'
    # 输入sdf路径
    sdf_path = '1_1w.sdf'
    # 注意: 仅在第一次操作的时候转化为xlsx。
    '''
        # 格式一: excel文件和此程序在同一个文件夹下，直接输入名字即可。
        # path_name='1.csv'
        # 格式二写法： 注意是'/',别写错了。 (没有做优化了，可以优化修改路径的)
        # path_name='C:/All_Softwares/Develop_Tools/Idea_Project/Src/Python_project/Python_Task/Task/2022_April/pubchem_code/1.csv'
        ConvertToExcel(path_name=path_name)
    '''
    # 第一次运行 需要打开这个
    ConvertToExcel(path_name=path_name)

    # 这个是保存你想要存储转化好的mol文件存储文件夹的名字
    filename = '1_1w'

    # 默认路径
    savepath = './ALL_Mol/{filename}'.format(filename=filename)
    # default_flag=True代表创建默认路径
    creat_save_path(filename, load_path='./ALL_Mol', default_flag=True)
    '''
    # 自定义路径： 格式按照我这种类似的写
    load_path='E:\\DATA'
    save_path= creat_save_path(filename, load_path=load_path)
    '''
    # 打开自定义路径    default_flag参数不用写进去，我跳过验证了，按照下面的格式写就行
    # load_path='E:\\DATA'
    # save_path= creat_save_path(filename, load_path=load_path)

    sleep(2)
    # 读取excel文件
    excel_path_f = path_name.split('.')[0] + '.xlsx'
    # 获取一个存储对应次序的cas号，为提高代码简洁度，这里仅采用列表方式，一次性存储。
    # turn_flag=True 默认开启，代表把cas存储到excel中，如果是第二次运行，则设置 turn_flag=False
    temp_data = get_cas_list(excel_path=r'{excel_path_f}'.format(excel_path_f=excel_path_f), turn_flag=True)
    cas_list = temp_data[0]
    cid_list = temp_data[1]
    #  将数据存储到cid_dic字典里面
    cid_dic = get_cid_dic(cid_list, cas_list)
    # 用来标记存储过的cas
    saved_caslist=[]

    # 读取sdf的数据全部存储为列表了 写入sdf的路径
    sdf_file = Read_SDF(read_path=str(sdf_path))

    current_index = 0

    temp_list = []
    # flag用来标记是否处在多余部分
    flag = False
    for index in tqdm(range(0, len(sdf_file))):
        if (flag == False and sdf_file[index] != 'M  END\n'):
            temp_list.append(sdf_file[index])
        elif (sdf_file[index] == 'M  END\n'):
            # 一个sdf的内容全部拿到，临时存储至temp_list
            temp_list.append(sdf_file[index])
            # 判断sdf的cid 和存储的cid是否对的上 去除换行的
            sdf_cid = temp_list[0].strip()

            # 根据cas号开始命名存储文件
            convert_mol(savepath, temp_list, current_index, sdf_cid)
            # convert_mol(savepath=str(savepath),temp_list=temp_list, current_index=current_index)
            current_index += 1
            # 清空临时列表
            temp_list = []
            # 这个时候进入多余部分，修改flag
            flag = True
            # 当达到cas_list的长度的时候，程序结束
            if (current_index == len(cas_list)):
                print("*********数据转化结束。感谢您的使用！")
                t2 = time.time()
                t = round(t2 - t1, 2)
                print('消耗时间{t}(s)'.format(t=t))
                break
        elif (sdf_file[index] == '$$$$\n'):
            flag = False
