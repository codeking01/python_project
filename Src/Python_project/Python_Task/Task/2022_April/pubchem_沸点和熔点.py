# https://pubchem.ncbi.nlm.nih.gov/#query=57-27-2
import json
import time
import os

from lxml import etree

import requests
import openpyxl
from openpyxl import *
from tqdm import tqdm

wbk = Workbook()
ws = wbk.active
ws.title = 'cas_熔点'
ws['A1'] = 'cas'
ws['B1'] = '熔点'
# 用来操作相同数据

wb = openpyxl.load_workbook('熔点CAS.xlsx')
sheet = wb['data']
rows = sheet.max_row
table_rows = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',
              'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36',
    'Content-Type': 'application/json; charset=utf-8',
}
for z in tqdm(range(1153, rows + 1)):
    try:
        cas = sheet.cell(row=z, column=2).value
        cas = cas.replace('_', '-')
        # 获取起始地址
        # url = 'https://pubchem.ncbi.nlm.nih.gov/#query=57-27-2'
        # cas = '57-27-2'
        url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
            cas) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
        response_1 = requests.get(url=url, headers=headers, timeout=220)
        response_1.encoding = 'utf-8'
        json_content = response_1.text
        obj = json.loads(json_content)
        # 获取查询页面
        obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
        url_boil = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{0}/JSON/'.format(obj_cid)
        response = requests.get(url=url_boil, headers=headers, timeout=220)
        response.encoding = 'utf-8'
        obj = json.loads(response.text)
        # Boiling_point_pre = obj['Record']['Section'][3]['Section']
        Boiling_point_pre = obj['Record']['Section']
        point = '无'
        tag = False  # 标识沸点是否存在
        Boiling_point = []
        # 拿到头部
        for j in Boiling_point_pre:
            if (j['TOCHeading'] == 'Chemical and Physical Properties'):
                Boiling_point = j['Section']
                break
        # 获取沸点
        for x in Boiling_point:
            if (x['TOCHeading'] == 'Experimental Properties'):
                for k in x['Section']:
                    if (k['TOCHeading'] == 'Melting Point'):
                        tag = True
                        count = 0
                        # print(k['Information'])
                        for l in k['Information']:
                            if ('StringWithMarkup' in l['Value']):
                                point = l['Value']['StringWithMarkup'][0]['String']
                                print(point)
                            else:
                                point_con = str(l['Value']['Number'][0])
                                # print(point_con)
                                Unit = l['Value']['Unit']
                                # print(Unit)
                                point = point_con + Unit
                            print('第{0}条的熔点'.format(z), point)
                            ws['A' + str(z)].value = cas
                            ws[table_rows[count] + str(z)].value = point
                            count += 1
                        break
                break
        if (tag == False):
            ws['A' + str(z)].value = cas
            ws['B' + str(z)].value = point

        wbk.save('cas_pubchem熔点1.xlsx')
        print(cas, '第{0}处理成功*****'.format(z))
    except:

        ws['A' + str(z)].value = cas
        ws['B' + str(z)].value = '数据可能有问题,请手动查询....'
        wbk.save('cas_pubchem熔点1.xlsx')
        print(cas, '第{0}处理失败......'.format(z))
