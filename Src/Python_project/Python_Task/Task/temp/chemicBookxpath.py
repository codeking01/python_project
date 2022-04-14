import requests
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from lxml import etree

if __name__ == '__main__':
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:98.0) Gecko/20100101 Firefox/98.0"
    }
    # 这个是第一次创建的时候写的 这个可以自己手动创建
    wb = Workbook()
    ws = wb.active
    ws.title = 'chemicbook数据'
    ws['A1'] = '原始CAS'
    ws['B1'] = 'CAS'
    ws['C1'] = 'Chemical_Name'
    ws['D1'] = 'MF'
    wb.save('杨喜年的数据5.xlsx')

    wb = load_workbook('杨喜年的数据5.xlsx')
    # 取第一张表
    sheetnames = wb.sheetnames
    # 获取第一张表
    ws = wb[sheetnames[0]]
    for page in range(0, 610):
        all_cas = []
        chemic_name = []
        all_mf = []
        url = 'https://www.chemicalbook.com/ProductCASList_16_{page}00_EN.htm'.format(page=page)
        resp = requests.get(url=url, headers=headers)
        page_content = resp.text
        tree = etree.HTML(page_content)
        # 2到102 一共100个数据
        for i in range(2, 102):
            # 获取所有的cas号
            all_cas.append(tree.xpath('//table//tr[{i}]/td[3]/a/text()'.format(i=i))[0])
            # 获取所有的chemic_name
            chemic_name.append(tree.xpath('//table//tr[{i}]/td[2]/a/text()'.format(i=i))[0])
            # 获取所有的all_mf
            # 看看mf解析是否有数据,没有数据就给 空的存进列表
            if(len(tree.xpath('//table//tr[{i}]/td[4]/span/text()'.format(i=i)))==0):
                all_mf.append('')
            else:
                all_mf.append(tree.xpath('//table//tr[{i}]/td[4]/span/text()'.format(i=i))[0])

        # CAS //table//tr[2+]/td[3]/a/text()
        # Chemical_Name //table//tr[2+]/td[2]/a/text()
        # MF //table//tr[2+]/td[4]/span/text()
        # 把每一页的内容存到excle里面
        for i in tqdm(range(len(all_cas))):
            try:
                ws['A' + str(page * 100 + i + 2)].value = all_cas[i]
                ws['B' + str(page * 100 + i + 2)].value = all_cas[i].replace('-', '_')
                ws['C' + str(page * 100 + i + 2)].value = chemic_name[i]
                ws['D' + str(page * 100 + i + 2)].value = all_mf[i]
            except Exception as e:
                print('第{page}页失败,reason'.format(page=page), e)
        print('第{page}页存储成功'.format(page=page))
        if (page%200==0):
            wb.save('杨喜年的数据5.xlsx')
    print("数据全部采集成功！")     
    wb.save('杨喜年的数据5.xlsx')

'''
from lxml import  etree
tree=etree.HTML(page_content)
content=tree.xpath('//td[2]/a/text()')
for i in content:
    if i not in chemic_name:
        print(i)
'''
