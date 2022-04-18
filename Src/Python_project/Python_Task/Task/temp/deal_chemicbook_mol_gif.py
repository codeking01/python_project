#  todo 下面的 代码重新移植到新的py
# 再次处理cas
# todo 加入数据库
# new_url = 'https://www.chemicalbook.com/Search.aspx?_s=&keyword={cas}'.format(cas=cas)
#  todo 考虑加入代理去访问
# response = requests.get(new_url)


# 处理最开的excel的数据
import openpyxl
from tqdm import tqdm


# todo 把读文件和读取excle的方法封装成一个utils的类 以后直接复用
def Predeal_Data():
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['cas']
    cas_list=[]
    deal_cas_list=[]

    # 读取excel中的最大行数
    rows = sheet.max_row
    # 处理所有数据
    for sheet_row in tqdm(range(rows + 1)):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=sheet_row + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
        except:
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
            print(cas, deal_cas)
    print("数据全部加载成功...")
    return rows


if __name__ == '__main__':

    new_url = 'https://www.chemicalbook.com/Search.aspx?_s=&keyword={cas}'.format(cas=cas)