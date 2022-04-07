import requests
import re

from openpyxl import Workbook
from tqdm import tqdm

if __name__ == '__main__':
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:98.0) Gecko/20100101 Firefox/98.0"
    }
    wb = Workbook()
    ws = wb.active
    ws.title = 'chemicbook数据'
    ws['A1'] = 'CAS'
    ws['B1'] = 'MF'
    for page in range(0, 367):
        caslist = []
        mflist = []
        url = 'https://www.chemicalbook.com/ProductCASList_15_{page}00_EN.htm'.format(page=page)
        resp = requests.get(url=url, headers=headers)
        page_content = resp.text
        obj = re.compile(r'<td style="width:200px;">(?P<name>.*?)</td>', re.S)
        result = obj.finditer(page_content)
        print(result)
        cas_pattern = re.compile(r'<a.*">(?P<cas_name>.*?)</a>')
        mf_pattern = re.compile(r'<span.*>(.*)</span>')
        flag = True
        # 每一页都是100个
        num = 100

        for it in result:
            img_contens = it.group("name").strip()
            # print('初始的',img_contens)
            a = img_contens.replace('..\\', '')
            if (flag == True):
                all_cas = cas_pattern.search(a)
                cascontents = all_cas[1].replace('-','_')
                caslist.append(cascontents)
                # print('cas为:', all_cas.group(1))
            else:
                mf = mf_pattern.search(a)
                if (mf.group(1) == ''):
                    mfcontents = '无'
                    mflist.append(mfcontents)
                    # print('Mf 无')
                else:
                    mfcontents = mf.group(1)
                    mflist.append(mfcontents)
                    # print('mf为:',mf.group(1))
            # 设置flag反转
            flag = bool(1 - flag)

        # 把每一页的内容存到excle里面
        for i in tqdm(range(len(caslist))):
            ws['A' + str(page * 100 + i + 2)].value = caslist[i]
            ws['B' + str(page * 100 + i + 2)].value = mflist[i]
        print('第{page}页存储成功'.format(page=page))
        wb.save('YAcas数据.xlsx')
    # 处理转义字符
    # img_contens=img_contens.encode('raw_unicode_escape').decode('utf8')

    # # print('a:',a)
    # all_cas=cas_pattern.search(a)
    # print('cas为:',all_cas[1])
    #
    # mf=mf_pattern.search(a)
    # if (mf==None):
    #     print('Mf 没有')
    # else:
    #     print('mf为:',mf)
    # print('匹配的',all_cas[])
    # print('匹配的',all_cas)

    # print(all_cas[1])
    # try:
    #     print('匹配为：',all_cas[0])
    # except:
