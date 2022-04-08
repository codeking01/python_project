import requests
import re

from openpyxl import Workbook, load_workbook
from tqdm import tqdm


# 去除所有的\n\t\r\f
def del_escape(old_list):
    for i in range(len(old_list)):
        if '\n' in old_list[i]:
            old_list[i] = old_list[i].replace('\n', ' ')
        if '\t' in old_list[i]:
            old_list[i] = old_list[i].replace('\t', ' ')
        if '\f' in old_list[i]:
            old_list[i] = old_list[i].replace('\f', '')
        if '\r' in old_list[i]:
            old_list[i] = old_list[i].replace('\r', '')
    return old_list


if __name__ == '__main__':
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:98.0) Gecko/20100101 Firefox/98.0"
    }
    # 这个是第一次创建的时候写的
    # wb = Workbook()
    # ws = wb.active
    # ws.title = 'chemicbook数据'
    # ws['A1'] = '原始CAS'
    # ws['B1'] = 'CAS'
    # ws['C1'] = 'Chemical_Name'
    # ws['D1'] = 'MF'
    # wb.save('杨喜年的数据.xlsx')

    wb = load_workbook('杨喜年的数据.xlsx')
    # 取第一张表
    sheetnames = wb.sheetnames
    # 获取第一张表
    ws = wb[sheetnames[0]]
    for page in range(184, 367):
        url = 'https://www.chemicalbook.com/ProductCASList_15_{page}00_EN.htm'.format(page=page)
        resp = requests.get(url=url, headers=headers)
        page_content = resp.text
        # 去除'\x00'转义字符
        page_content = re.sub(u'\x00', '', page_content)
        page_content = re.sub(u'\x0b', '', page_content)
        # todo 解决转义字符\n
        pattern_name = re.compile(
            r'<a class=blue onclick="blur.*" href="/ProductChemicalProperties.*htm">(.*\s*.*)</a>' or r'<a class=blue onclick="blur.*" href="/ProductChemicalProperties.*htm">(.*\\*.*)</a>')
        # 获取所有的name
        chemic_name = pattern_name.findall(page_content)
        chemic_name = del_escape(chemic_name)
        # 获取所有的cas
        pattern_cas = re.compile(
            r'<a class=blue onclick="blur.*href="/ProdSupplier.*>(.*?\s*.*)</a>' or r'<a class=blue onclick="blur.*href="/ProdSupplier.*>(.*?\\*.*)</a>')
        all_cas = pattern_cas.findall(page_content)
        all_cas = del_escape(all_cas)
        # 获取所有的mf
        pattern_mf = re.compile(
            r'<span id="ContentPlaceHolder1_ProductClassDetail.*">(.*\s*.*)</span>' or r'<span id="ContentPlaceHolder1_ProductClassDetail.*">(.*\\*.*)</span>')
        all_mf = pattern_mf.findall(page_content)
        all_mf = del_escape(all_mf)

        # 把每一页的内容存到excle里面
        for i in tqdm(range(len(all_cas))):
            try:
                ws['A' + str(page * 100 + i + 2)].value = all_cas[i]
                ws['B' + str(page * 100 + i + 2)].value = all_cas[i].replace('-', '_')
                ws['C' + str(page * 100 + i + 2)].value = chemic_name[i]
                ws['D' + str(page * 100 + i + 2)].value = all_mf[i]
            except Exception as e:
                print('第{page}页失败,reason'.format(page=page), e)
        print('第{page}页存储成功'.format(page=page))
        wb.save('杨喜年的数据.xlsx')


'''
from lxml import  etree
tree=etree.HTML(page_content)
content=tree.xpath('//td[2]/a/text()')
for i in content:
    if i not in chemic_name:
        print(i)
'''