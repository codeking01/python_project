'''
@author king_xiong
@date 2021-12-16 20:02
'''
import os
import time
import urllib.request
import openpyxl
import requests
import json
import winsound
# 页面地址
# https://pubchem.ncbi.nlm.nih.gov/compound/5288826
# 下载地址
# https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/5288826/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_5288826
from tqdm import tqdm

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
}
class FontColor:
    OKBULE = '\033[94m'
    OKGREEN = '\033[92m'
    WARN = '\033[93m'
    FAIL = '\033[91m'
f = FontColor()

# proxies = {
#     # 这个位置加代理ip
# }


# import pymysql
# # 打开数据库连接
# db = pymysql.connect(host="localhost", user="root", password="123456", database="417database")
# # 使用cursor（）方法获取游标
# cursor=db.cursor()
# # 执行SQL语句
# insert_sql = "insert into test01(id, name, age) values(%s, %s, %s)"
# parm=(4,"是的",2)
# # 执行sql语句
# cursor.execute(insert_sql,parm)
# # 提交到数据库执行
# db.commit()
# cursor.execute("select * from test01")
# # 查看表里所有数据
# data = cursor.fetchall()
# print(data)


def IS_Right(func):
    def wrapper(*args):
        flag=func(*args)
        if(flag!="NULL"):
            obj_cid = str(flag)
            return obj_cid
        else:
            return "erro"
    return wrapper

@IS_Right
def get_cid(deal_cas):
    try:
        # 下面这个是为了获取需要链接的cid号
        url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
            deal_cas) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
        response = requests.get(url=url, headers=headers)
        response.encoding = 'utf-8'
        # 下面这个json数据可以采集到
        json_content = response.text
        # print(page_content)
        # ***可以直接读取json，就不存进去了
        obj = json.loads(json_content)
        # 找到json中相应的数据
        obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
        return obj_cid
    except:
        return "NULL"


def Is_Download(func):
    def wrapper(*args):
        if(obj_cid!="erro"):
            func(*args)
        else:
            winsound.Beep(600, 1000)
            print(f.WARN+'{cas}的cid获取不到,的mol和png文件下载失败..'.format(cas=cas))
            time.sleep(2)
    return wrapper


@Is_Download
def download_mols(obj_cid, cas):
    # 获取下载链接https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/5288826/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_5288826
    url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/{obj_cid}'.format(
        obj_cid=obj_cid) + '/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_{obj_cid}'.format(
        obj_cid=obj_cid)
    urllib.request.urlretrieve(url, 'D:/Task_Datas/ALL_MOLS/' + str(cas) + '.sdf')
    print('第{0}条数据的mol文件处理成功'.format(t))

@Is_Download
def download_pngs(obj_cid):
    # 需要下载的图片地址  https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid=5288826
    pic_url = 'https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid={cid}'.format(cid=obj_cid)
    urllib.request.urlretrieve(pic_url, 'D:/Task_Datas/ALL_PNGS/' + str(cas) + '.png')
    print('第{0}条数据的png文件处理成功'.format(t))


# 最开始处理excle里面的数据
def Predeal_Data():
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['cas']
    # 读取excel中的最大行数
    rows = sheet.max_row

    # 处理所有数据
    for t in tqdm(range(rows + 1)):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=t + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
        except:
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
            print(cas,deal_cas)
    print("数据全部加载成功...")
    return rows


if __name__ == '__main__':
    '''
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['CAS']
    # 读取excel中的最大行数
    rows = sheet.max_row
    # 创建一个存储的cas初始值列表
    cas_list=[]
    # 创建一个存储的cas修改后值列表
    deal_cas_list=[]
    # 处理所有数据
    for t in range(rows + 1):
        # 读取cas号，从第二行开始
        cas = sheet.cell(row=t+2, column=1).value
        deal_cas=cas.replace('_','-')
        cas_list.append(str(cas))
        deal_cas_list.append(str(deal_cas))
        print("数据全部加载成功")
    '''

    # 创建一个存储的cas初始值列表
    cas_list = []
    # 创建一个存储的cas修改后值列表
    deal_cas_list = []

    # 获取最大行数
    rows = Predeal_Data()
    # 一次到位的话就写成 rows+1

    for t in tqdm(range(4586,5000)):
        try:
            # 读取列表的数据值
            cas = str(cas_list[t])
            deal_cas = str(deal_cas_list[t])
            '''
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=t + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            '''
            '''
            # 下面这个是为了获取需要链接的cid号
            url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
                deal_cas) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
            response = requests.get(url=url, headers=headers)
            response.encoding = 'utf-8'
            # 下面这个json数据可以采集到
            json_content = response.text
            # print(page_content)
            # ***可以直接读取json，就不存进去了
            obj = json.loads(json_content)
            # 找到json中相应的数据
            obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
            '''

            # 获取所需的cid
            obj_cid = get_cid(deal_cas)

            # 下载mol文件(sdf的)
            download_mols(obj_cid, cas)
            '''
            # 获取下载链接https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/5288826/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_5288826
            url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/{obj_cid}'.format(obj_cid=obj_cid)+'/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_{obj_cid}'.format(obj_cid=obj_cid)
            urllib.request.urlretrieve(url, 'D:/Task_Datas/ALL_MOLS' + str(cas) + '.sdf')
            print('第{0}条数据处理成功'.format(t))
            '''

            # 下载png文件
            download_pngs(obj_cid)
            '''
            # 需要下载的图片地址  https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid=5288826
            pic_url='https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid={cid}'.format(cid=obj_cid)
            urllib.request.urlretrieve(pic_url, 'D:/Task_Datas/ALL_PNGS' + str(cas) + '.png')
            '''

            if (t % 2000 == 0 and t != 0):
                print('休息500s,不要着急，小心封号！')
                time.sleep(500)
        except:
            # 播放声音3s
            winsound.Beep(500, 1000)
            print(f.WARN+'*********第{0}条数据处理失败....'.format(t))
            print(f.WARN+'---------失败的cas为' + str(deal_cas))
