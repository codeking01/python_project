import os
import time
import urllib.request

import openpyxl
import requests
from lxml import etree
class FontColor:
    OKBULE = '\033[94m'
    OKGREEN = '\033[92m'
    WARN = '\033[93m'
    FAIL = '\033[91m'
f = FontColor()
# print("\033[0;31m%s\033[0m" % "输出红色字符")

# print(f.WARN  +'你好')
from openpyxl import *

# https://www.chemicalbook.com/Search.aspx?keyword=1004-6-6
# https://www.chemicalbook.com/CAS/MOL/100-02-7.mol

# import pymysql
# # 打开数据库连接
# db = pymysql.connect(host="localhost", user="root", password="123456", database="417database")
# # 使用cursor（）方法获取游标
# cursor=db.cursor()
# # 执行SQL语句
# insert_sql = "insert into test01(id, name, age) values(%s, %s, %s)"
# parm=(4,"是的",2)
# # 执行sql语句
# cursor.execute(insert_sql,parm)
# # 提交到数据库执行
# db.commit()
# cursor.execute("select * from test01")
# # 查看表里所有数据
# data = cursor.fetchall()
# print(data)
if __name__ == '__main__':
    # 用来记录连续失败次数
    coin =0
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['CAS']
    # 读取excel中的最大行数
    rows = sheet.max_row
    for t in range(15773,rows+1):
        # 11904 13032（还要+1=13033）
        # 13501
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=t + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            # https://www.chemicalbook.com/CAS/20200119/MOL/52897-01-5.mol
            # https://www.chemicalbook.com/CAS/20200611/MOL/52897-03-7.mol
            # xpath 链接  //tr/td[1]//div/a[1]/@href
            # 获取下载链接代码:

            # pre_url ='https://www.chemicalbook.com/Search.aspx?keyword='+str(deal_cas)
            # content = requests.get(pre_url).text
            # tree = etree.HTML(content)
            # url='https://www.chemicalbook.com/'+tree.xpath("//tr/td[1]//div/a[1]/@href")[0]
            # print(url)

            # 下面这个下载链接写的有问题，上面重新使用xpath获取地址
            url = 'https://www.chemicalbook.com/CAS/MOL/{0}.mol'.format(deal_cas)
            urllib.request.urlretrieve(url, './mol_files_chembook/' + str(cas) + '.mol')
            # 爬取成功，请空失败次数
            coin=0
            print('第{0}条数据处理成功'.format(t))
            time.sleep(1)
            # if(t%1000==0 and t!=0 ):
            #     time.sleep(300)
        except:
            # 采取第二中2下载的方法：用xpath取获取一些隐藏数据，完善下载链接（但是访问次数容易收到限制，需要经常验证）
            pre_url ='https://www.chemicalbook.com/Search.aspx?keyword='+str(deal_cas)
            content = requests.get(pre_url).text
            tree = etree.HTML(content)
            # 处理没有mol的情况
            if(len(tree.xpath("//tr/td[1]//div/a[1]/@href"))==0):
                print('*********第{0}条数据处理失败....'.format(t))
                print('---------失败的cas为 '+str(deal_cas),pre_url)
            else:
                url='https://www.chemicalbook.com/'+tree.xpath("//tr/td[1]//div/a[1]/@href")[0]
                if(requests.get(url=url).status_code==200):
                    urllib.request.urlretrieve(url, './mol_files_chembook/' + str(cas) + '.mol')
                    print('第{0}条数据处理成功'.format(t),'第二种方法的url为  '+url)
                # 第二种下载方法失败记录连续失败次数
                else:
                    # 记录连续失败次数
                    coin+=1
                    # 如果连续失败4次以上，可能ip被封了，检查一下
                    if(coin> 4):
                        # 失败次数过多播放音乐，提醒检查网址
                        file_path = "C:\\Users\\king\Downloads\Music\错位时空.mp3"
                        os.startfile(file_path)
                        time.sleep(10)
                    print(f.WARN+'*********第{0}条数据处理失败....'.format(t))
                    print(f.WARN+'---------失败的cas为 '+str(deal_cas),url)
    print('爬取完成了！！')