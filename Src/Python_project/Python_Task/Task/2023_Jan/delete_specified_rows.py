# author: code_king
# time: 2023/1/10 21:28
# file: delete_specified_rows.py
import os

import openpyxl
from tqdm import tqdm


def get_file_name_list(gjf=None):
    """
    :param gjf: 文件夹路径
    :return: 所有结构图的名字列表
    """
    # 读取指定路径所有的文件名
    path = r"%s" % gjf
    file_list = os.listdir(path)
    # 读取以gif 结尾的文件
    names_list = set()
    for i in file_list:
        try:
            if i.split('.')[1] == "gjf":
                names_list.add(i.split('.')[0])
        except Exception as e:
            print("失败原因：", e)
            print("应该是文件名的问题..")
    # print(len(names_list))
    return list(names_list)


def deal_excel_data(excel_path=None, table_name="sheet1", start_row=1, order_col=1, deal_data=None):
    """
    :param excel_path:
    :param table_name:
    :param start_row:
    :param order_col:
    :param deal_data:
    :return: 处理好excel，在需要删除的位置，提前添加 ”$“
    """
    if excel_path != None:
        excel_path = r"%s" % excel_path
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[f'{table_name}']
    # 读取excel中的最大行数
    rows = sheet.max_row
    datas = []
    print(f"删除前最大行数：{rows}")
    # 读取从start_row开始的所有行的 第order_col列 这个地方需要倒着读
    for i in tqdm(range(rows, start_row, -1)):
        current_data = sheet.cell(row=i, column=order_col).value
        if current_data not in names_list:
            # 做一个要被删除的标记 "$"
            sheet.cell(row=i, column=order_col).value = "$"
            # print(current_data)
    wb.save(excel_path)

    # 测试
    # 删除"$" 这一行 删除行和清空行数据不同，删除行后下面的行会往上移，所以这里行号排序倒着删就不会出现顺序删除时，部分行没有被删掉
    # for i in tqdm(range(rows, start_row, -1)):
    #     current_data = sheet.cell(row=i, column=order_col).value
    #     if current_data == "$":
    #         # 数据会上移
    #         sheet.delete_rows(i)
    # print(f"删除后最大行数：{sheet.max_row}")
    # wb.save(excel_path)


def delete_excel_data(excel_path=None, table_name="sheet1", start_row=1, order_col=1):
    """ 为了解耦和以后复用，分开写了
    :param excel_path:
    :param table_name:
    :param start_row:
    :param order_col: 指定的列
    :return: 删除好的excel数据
    """
    if excel_path != None:
        excel_path = r"%s" % excel_path
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[f'{table_name}']
    # 读取excel中的最大行数
    rows = sheet.max_row
    # 删除"$" 这一行 删除行和清空行数据不同，删除行后下面的行会往上移，所以这里行号排序倒着删就不会出现顺序删除时，部分行没有被删掉
    for i in tqdm(range(rows, start_row, -1)):
        current_data = sheet.cell(row=i, column=order_col).value
        if current_data == "$":
            # 数据会上移
            sheet.delete_rows(i)
    print(f"删除后最大行数：{sheet.max_row}")
    wb.save(excel_path)


if __name__ == "__main__":
    # 使用进度条 tqdm
    names_list = get_file_name_list(gjf=r"C:\Users\king\Desktop\hxy_task1_10\structure\gjf")
    # 读取excel
    excel_path = r"C:\Users\king\Desktop\hxy_task1_10\design1.xlsx"
    deal_excel_data(excel_path=excel_path, table_name="data", start_row=2, order_col=1, deal_data=names_list)
    delete_excel_data(excel_path=excel_path, table_name="data", start_row=2, order_col=1)
