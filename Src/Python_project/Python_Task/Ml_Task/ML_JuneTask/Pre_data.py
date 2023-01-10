# author: code_king
# time: 2022/11/24 22:37 
# file: Pre_data.py
"""
还是5%异常来处理
前面阶段的输入和输出是否都是下一个阶段的输入’和‘前面阶段的输入是下一个阶段的输入’两种方式都尝试一下
"""
import copy
import os
import warnings

import joblib
import numpy as np
import pandas as pd
from numpy import unique

from ML_JuneTask.DimensionalityReduction import DataDelete
from ML_JuneTask.NewPreDeal_Tools import Del_deletion_data, get_final_useablecols, all_ydata, \
    delAndGetCols, get_merge_tabledata, get_former_Ydata, convert_to_num, deal_verify_ydata, Deal_sorted_Ydata
from ML_JuneTask.build_images.init_build_images import gen_images
from utils.plot_error_distribution import write_images

# warnings.filterwarnings(action='always', category=UserWarning)
warnings.filterwarnings(action='ignore')
# 这个pandas处理数据效果不太好 建议用numpy
thedata = pd.read_excel(io=r'data_original.xlsx', sheet_name='8000D数据标准化-建模用')
Preddata = pd.read_excel(io=r'data_original.xlsx', sheet_name='8000D数据标准化-预测用')
# 用来查看数据结果的excel对象
from openpyxl import load_workbook

filename = 'data_original.xlsx'
data = load_workbook(f'{filename}')
sheetnames = data.sheetnames
# create_sheet
if len(sheetnames) < 3:
    sheetnames.append('预测结果')
    data.create_sheet(sheetnames[2], len(sheetnames))
    # 赋值sheet
    # sheet=data[sheetnames[0]]
    # content=data.copy_worksheet(sheet)
    data.save(f'{filename}')
# 预测结果的excel表
table = data[sheetnames[2]]
sheetnames = data.sheetnames
table = data[sheetnames[2]]

# 获取数据表对象(建模和与预测的数据)
thedata = np.array(thedata)
Preddata = np.array(Preddata)


# 写入excel文件
def write_to_excel(start_index, table, current_column, content):
    # 标签
    for i in range(len(content)):
        table.cell(start_index, current_column).value = content[i]
        start_index += 1


def save_model(model, model_name, X_train, X_test, y_train, y_test, train_numbers):
    """
    :param model: 需要训练的模型
    :param X_train:
    :param X_test:
    :param y_train:
    :param y_test:
    :return:
    """
    # 训练次数
    if not os.path.exists(f"models/{model_name}"):
        os.makedirs(name=f"models/{model_name}", exist_ok=True)
    acc_dic = {}
    for i in range(11):
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        # 取最小的
        model_train_acc = min(model_train_acc, model_acc)
        acc_dic.update({f"model[{i}]": model_train_acc})
        # 模型保存
        joblib.dump(model, f'models/{model_name}/model[{i}].pkl')
    for i in range(11, train_numbers):
        # temp_keys 是键的按照值从小到打排序的列表 [model[1],model[5],...]
        temp_keys = sorted(acc_dic, reverse=True)
        model.fit(X_train, y_train.astype(float))
        # 训练集准确率
        model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
        # print('Rfc_train_acc准确率：', model_train_acc)
        model_pred = model.predict(X_test)
        model_acc = accuracy_score(y_test.astype(float), model_pred)
        model_train_acc = min(model_train_acc, model_acc)
        # 替换一个最小的模型
        for key_item_index in range(0, len(temp_keys)):
            if key_item_index != len(temp_keys) - 1:
                if acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc < acc_dic[
                    f"{temp_keys[key_item_index + 1]}"]:
                    acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                    joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')
            elif acc_dic[f"{temp_keys[key_item_index]}"] < model_train_acc:
                acc_dic.update({f"{temp_keys[key_item_index]}": model_train_acc})
                joblib.dump(model, f'models/{model_name}/{temp_keys[key_item_index]}.pkl')

    # model1 = joblib.load(filename="filename.pkl")


def get_pred_data(model_name, X_verify_data):
    """
    :param model_name: 模型名字
    :return: pred_data
    """
    # 预测数据
    all_models = os.listdir(f"models/{model_name}")
    model_pred_list = []
    for i in all_models:
        cursor_model = joblib.load(filename=f"models/{model_name}/{i}")
        model_pred_list.append(cursor_model.predict(X_verify_data))
    model_pred_list = np.array(model_pred_list)
    temp_pred_list = []
    # 循环遍历列
    for pred_list_index in range(0, model_pred_list.shape[1]):
        # 预测结果统一,那个数多就等于哪个
        zero_count = np.where(model_pred_list[:, pred_list_index] == 0)[0].shape[0]
        negative_count = np.where(model_pred_list[:, pred_list_index] == -1)[0].shape[0]
        positive_count = np.where(model_pred_list[:, pred_list_index] == 1)[0].shape[0]
        final_dic = {}
        final_dic.update({"0": zero_count})
        final_dic.update({"-1": negative_count})
        final_dic.update({"1": positive_count})
        # 排序，第一个数预测的最多
        final_result = -1
        if final_dic[f"{final_result}"] < final_dic["0"]:
            final_result = 0
        if final_dic[f"{final_result}"] < final_dic["1"]:
            final_result = 1
        temp_pred_list.append(final_result)
    pred_data = np.array(temp_pred_list)
    return pred_data


def get_train_test_acc(model, X_train, y_train, X_test, y_test):
    """
    :param model:
    :param X_train:
    :param y_train:
    :param X_test:
    :param y_test:
    :return: model_train_acc, model_acc
    """
    model_train_acc = accuracy_score(y_train.astype(float), model.predict(X_train))
    model_pred = model.predict(X_test)
    model_acc = accuracy_score(y_test.astype(float), model_pred)
    return model_train_acc, model_acc


if __name__ == '__main__':
    # 将表的内容 7个数据表 挨个处理
    # 获取源建模的X数据 传入的对象是获取的excel对象
    # 获取标题内容
    all_titles=thedata[1,:]
    FormerTwo_data, FormerThree_data, FormerFour_data, FormerFive_data, FormerSix_data = get_merge_tabledata(
        excel_data=thedata)
    # 获取预测的X数据
    Verify_TABLE_TWO, Verify_TABLE_THREE, Verify_TABLE_FOUR, Verify_TABLE_FIVE, Verify_TABLE_SIX = get_merge_tabledata(
        excel_data=Preddata)
    all_former_data = [FormerTwo_data, FormerThree_data, FormerFour_data, FormerFive_data, FormerSix_data]
    all_verify_data = [Verify_TABLE_TWO, Verify_TABLE_THREE, Verify_TABLE_FOUR, Verify_TABLE_FIVE, Verify_TABLE_SIX]
    # 依次遍历
    iter_length = len(all_former_data)
    for i in range(iter_length):
        # 统一关键字
        the_Former_data, the_Verify_TABLE_data = convert_to_num(all_former_data[i], all_verify_data[i])
        # 删除缺失值过多的列，并保存del_cols
        # tempFormerTwo_data = delAndGetCols(FormerTwo_data)
        base_Xdata, del_cols = delAndGetCols(the_Former_data)
        # 用del_cols删除验证集X不需要的列
        Verify_Xdata = np.delete(the_Verify_TABLE_data, del_cols, axis=1)
        # 降维处理,这个地方只从训练集判断相关性，然后统一降维·
        use_able_x_cols = DataDelete(base_Xdata, 0.80)
        base_Xdata = base_Xdata[:, use_able_x_cols]
        Verify_Xdata = Verify_Xdata[:, use_able_x_cols]
        # 合并数据一起作为训练集
        # all_train_x_data = np.r_[base_Xdata, Verify_Xdata]

        # base_Xdata,del_cols
        # %%
        # 获取源建模的数据Y_data
        Original_TableTwoYdata, Original_TableThreeYdata, Original_TableFiveYdata, Original_TableSixYdata, Original_TableSevenYdata = get_former_Ydata(
            thedata)
        # 获取预测数据的Y_data
        Predict_TableTwoYdata, Predict_TableThreeYdata, Predict_TableFiveYdata, Predict_TableSixYdata, Predict_TableSevenYdata = get_former_Ydata(
            Preddata)
        all_original_Ydata = [Original_TableTwoYdata, Original_TableThreeYdata, Original_TableFiveYdata,
                              Original_TableSixYdata, Original_TableSevenYdata]
        all_predict_Ydata = [Predict_TableTwoYdata, Predict_TableThreeYdata, Predict_TableFiveYdata,
                             Predict_TableSixYdata, Predict_TableSevenYdata]
        # %%
        # 将Y取出来，然后取出缺失值过多的列
        # 获取可用的列
        Original_TableYdata, Predict_TableYdata = all_original_Ydata[i], all_predict_Ydata[i]
        final_cols = get_final_useablecols(Original_TableYdata, Predict_TableYdata)
        # final_cols
        # %%
        # 获取可以用的Y_data
        Original_YdataList, Pred_YdataList = all_ydata(final_cols, Original_TableYdata, Predict_TableYdata)
        # 合并数据，训练的时候一起
        # all_train_y_data = np.r_[Original_YdataList, Pred_YdataList]
        # %%
        # 挨个遍历，取出有用的Y_data
        # 传入起始的坐标 24,67,108,145,179
        # 传入起始的坐标新表 22,31,50,71,88
        colum_list = [22, 31, 50, 71, 88]
        the_column = colum_list[i]
        start = 0
        for index in final_cols:
            # 建模的Y_data
            Modeling_Y_data = Original_YdataList[start]
            # all_train_y_data全部用来训练
            # Modeling_Y_data = Original_YdataList[:int(Pred_YdataList.shape[0] / 2)+1, :][start]
            # 预测的Y_data，取一半测试，一半预测
            Pred_Y_data = Pred_YdataList[start]
            # Pred_Y_data = Pred_YdataList[int(Pred_YdataList.shape[0] / 2) :, :][start]
            start += 1
            # 处理建模的X_data,Y_data
            temp_OriginalXdata = np.column_stack((base_Xdata, Modeling_Y_data))
            # 删除缺失的行
            temp_OriginalXdata = Del_deletion_data(temp_OriginalXdata, 0)
            # X_data,Y_data 是第一个表处理好的x,y
            X_data = temp_OriginalXdata[:, :-1]
            Y_data = temp_OriginalXdata[:, -1]
            # 处理验证的X_data,Y_data
            # 合并数据
            temp_verifydata = np.column_stack((Verify_Xdata, Pred_Y_data))
            # 删除缺失的行
            temp_verifydata = Del_deletion_data(temp_verifydata, 0)
            # 获取预测数据的X_data，Y_data
            X_verify_data = temp_verifydata[:, :-1]
            Y_verify_data = temp_verifydata[:, -1]
            print(list(set(np.where(np.isnan(temp_OriginalXdata.astype(float)) == True)[0].tolist())),
                  list(set(np.where(np.isnan(temp_verifydata.astype(float)) == True)[0].tolist())))

            # 作图
            # for x_item in range(0,X_data.shape[1]):
            #     # write_image(X_left=X_data[:,x_item], X_right=X_verify_data[:,x_item], names=f"表{i}的{x_item}列")
            #     write_images(X_left=X_data[:,x_item], X_right=X_verify_data[:,x_item], names=f"images/表{i}的{x_item}列.png")

            if 1 == 1:
                # 将建模Y_data分类(-1,0,1)，并且取出边界值
                # 保留原始的数据
                Y_data_original=copy.deepcopy(Y_data)
                Y_data, Y_data_boundsMin, Y_data_boundsMax = Deal_sorted_Ydata(data=Y_data)
                # 获取预测数据的Y_data Verify_Ydata = Pred_Y_data
                # 保留原始的数据
                Verify_Ydata_original = copy.deepcopy(Y_verify_data)
                Verify_Ydata = deal_verify_ydata(Y_verify_data, Y_data_boundsMin, Y_data_boundsMax)
                # 查看是否替换成功
                print([unique(Y_data), Y_data_boundsMin, Y_data_boundsMax], np.unique(Verify_Ydata))
                # 切分训练数据和测试数据

                ## 30%测试数据，70%训练数据，stratify=y表示训练数据和测试数据具有相同的类别比例 修改为0.25
                # X_train, X_test, y_train, y_test = train_test_split(X_data, Y_data, test_size=0.25, random_state=4,stratify=Y_data)
                # 手动划分数据集
                X_train, X_test, y_train, y_test = X_data, X_verify_data[:int(X_verify_data.shape[0] / 2) + 1,
                                                           :], Y_data, Y_verify_data[
                                                                       :int(Y_verify_data.shape[0] / 2) + 1]
                # 验证集
                X_verify_data = X_verify_data[int(X_verify_data.shape[0] / 2) + 1:, :]
                Y_verify_data = Y_verify_data[int(Y_verify_data.shape[0] / 2) + 1:]
                from sklearn.metrics import accuracy_score
                # 开始训练模型
                from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier

                # 设置训练次数
                train_numbers = 300
                # 获取模型  随机森林
                Rfc = RandomForestClassifier()
                model_name = "Rfc"
                save_model(Rfc, model_name, X_train, X_test, y_train, y_test, train_numbers)

                # 随便取一个作为训练集的结果
                current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
                Rfc_train_acc, Rfc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)

                Gbc = GradientBoostingClassifier()
                model_name = "Gbc"
                save_model(Gbc, model_name, X_train, X_test, y_train, y_test, train_numbers)
                current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
                Gbc_train_acc, Gbc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)
                from sklearn.svm import SVC

                Svc = SVC()
                model_name = "Svc"
                save_model(Svc, model_name, X_train, X_test, y_train, y_test, train_numbers)
                current_model = joblib.load(filename=f"models/{model_name}/model[0].pkl")
                Svc_train_acc, Svc_acc = get_train_test_acc(current_model, X_train, y_train, X_test, y_test)

                Rfc_pred = get_pred_data("Rfc", X_verify_data)
                Gbc_pred = get_pred_data("Gbc", X_verify_data)
                Svc_pred = get_pred_data("Svc", X_verify_data)

                verify_y1 = accuracy_score(Y_verify_data.astype(float), Rfc_pred)
                verify_y2 = accuracy_score(Y_verify_data.astype(float), Gbc_pred)
                verify_y3 = accuracy_score(Y_verify_data.astype(float), Svc_pred)

                vereify_dic = {}
                # 取分数最高的点保存
                max_scores = max(verify_y1, verify_y2, verify_y3)

                # gen_images(iter_left_data=None, iter_right_verify=None, files_name="final_results")
                write_images(data_left=Y_data_original, data_right=Verify_Ydata_original,names=f"final_results/{all_titles[the_column + index-1]}.png", score=max_scores)

                # 计算准确率
                print('验证集Rfc_acc准确率:', verify_y1)
                print('验证集Gbc_acc准确率:', verify_y2)
                print('验证集SVC准确率：', verify_y3)

                #  标签 都加20
                write_to_excel(start_index=5, table=table, current_column=1,
                               content=["训练集个数", "随机森林（Rfc）准确率", "梯度提升树（Gbc）准确率",
                                        "支持向量机(Svc)准确率", "测试集个数", "随机森林（Rfc）准确率",
                                        "梯度提升树（Gbc）准确率", "支持向量机(Svc)准确率", "验证集个数",
                                        "随机森林（Rfc）准确率", "梯度提升树（Gbc）准确率", "支持向量机(Svc)准确率", ])
                # 将训练集写入到excel中
                write_to_excel(start_index=5, table=table, current_column=the_column + index,
                               content=[X_train.shape[0], Rfc_train_acc, Gbc_train_acc, Svc_train_acc, X_test.shape[0],
                                        Rfc_acc, Gbc_acc, Svc_acc,
                                        X_verify_data.shape[0], verify_y1, verify_y2, verify_y3])

                data.save(filename)
    print("运行结束！")
