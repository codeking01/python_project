import json
import scrapy
import openpyxl
import winsound
from tqdm import tqdm
import requests
from Pubchemic_download.items import PubchemicDownloadItem

global cas, deal_cas, obj_cid, current_num, rows, headers
cas_list = []
deal_cas_list = []
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.20 Safari/537.36',
}


class FontColor:
    OKBULE = '\033[94m'
    OKGREEN = '\033[92m'
    WARN = '\033[93m'
    FAIL = '\033[91m'
f = FontColor()


# proxies = {
#     # 这个位置加代理ip
# }

# 判断cid是否合法
def IS_Right(func):
    def wrapper(*args):
        flag = func(*args)
        if (flag != "NULL"):
            obj_cid = str(flag)
            return obj_cid
        else:
            return "erro"
    return wrapper


@IS_Right
def get_cid(deal_cas):
    try:
        # 下面这个是为了获取需要链接的cid号
        url = 'https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22compound%22,%22where%22:{%22ands%22:[{%22cid%22:%22' + str(
            deal_cas) + '%22}]},%22order%22:[%22cid,asc%22],%22start%22:1,%22limit%22:10,%22width%22:1000000,%22listids%22:0}'
        response = requests.get(url=url, headers=headers)
        response.encoding = 'utf-8'
        # 下面这个json数据可以采集到
        json_content = response.text
        # print(page_content)
        # ***可以直接读取json，就不存进去了
        obj = json.loads(json_content)
        # 找到json中相应的数据
        obj_cid = obj["SDQOutputSet"][0]["rows"][0]["cid"]
        return obj_cid
    except:
        return "NULL"


# 处理最开的excel的数据
def Predeal_Data():
    wb = openpyxl.load_workbook('CAS_ALL.xlsx')
    sheet = wb['cas']
    # cas_list=[]
    # deal_cas_list=[]

    # 读取excel中的最大行数
    rows = sheet.max_row
    # 处理所有数据
    for sheet_row in tqdm(range(rows + 1)):
        try:
            # 读取cas号，从第二行开始
            cas = sheet.cell(row=sheet_row + 2, column=1).value
            deal_cas = cas.replace('_', '-')
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
        except:
            cas_list.append(str(cas))
            deal_cas_list.append(str(deal_cas))
            print(cas, deal_cas)
    print("数据全部加载成功...")
    return rows


def get_mol_url(obj_cid):
    url = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/CID/{obj_cid}'.format(
        obj_cid=obj_cid) + '/record/SDF/?record_type=2d&response_type=save&response_basename=Structure2D_CID_{obj_cid}'.format(
        obj_cid=obj_cid)
    return url


def get_png_url(obj_cid):
    pic_url = 'https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid={cid}'.format(cid=obj_cid)
    return pic_url


class PubchemicMolPngSpider(scrapy.Spider):
    # 爬虫的名字 运行爬虫文件时使用的名字
    name = 'Pubchemic_mol_png'
    # 爬虫允许的域名，在爬取的时候，如果不是此域名之下的
    allowed_domains = ['pubchem.ncbi.nlm.nih.gov']
    # 声明了爬虫的起始地址，可以写多个url，一般是一个
    start_urls = ['http://pubchem.ncbi.nlm.nih.gov/']  # 这个请求无效 我可以下面重写了请求的起始方法

    # def start_requests(self):
    #     pass

    def parse(self, response):
        # 获取总数据量
        rows = Predeal_Data()
        print("数据加载成功，开始爬~~~~~~~~")
        for current_num in tqdm(range(0, 1000)):
            # 获取所需的cid
            deal_cas = str(deal_cas_list[current_num])
            obj_cid = get_cid(deal_cas)
            if (obj_cid != "erro"):
                cas_name = cas_list[current_num]
                # 获取mol和png 的下载链接
                mol_url = get_mol_url(obj_cid)
                png_url = get_png_url(obj_cid)
                All_Data = PubchemicDownloadItem(mol_url=mol_url, png_url=png_url, current_num=current_num,cas=cas_name)
                yield All_Data
            else:
                mol_url = "NULL"
                png_url = "NULL"
                fail_cas = str(deal_cas_list[current_num])
                winsound.Beep(600, 1000)
                print(f.WARN + '{cas}的cid获取不到,的mol和png文件下载失败..'.format(cas=fail_cas))
                continue
                # All_Data = PubchemicDownloadItem(mol_url=mol_url, png_url=png_url,current_num=current_num)
                # yield All_Data
        print("数据全部加载成功...")
